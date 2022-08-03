## Excel Spreadsheet Automation
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# wb = xl.load_workbook("transactions.xlsx")   # Load the xlxs sheet called transactions.xlsx in the current directory and returns an object
# # print(type(wb))
# # sheet_1 = wb["Sheet1"] # Select Sheet1 in wb object (transactions.xlsx)
# # cell_1 = sheet_1['a1']  # Select A1 cell in sheet 1
# # cell_1 = sheet_1.cell(1, 1) # Another way to select cell A1 in sheet 1
#
# # print(cell_1.value) # Print the value in cell_1 object
# # print(type(cell_1.value))
#
# # print(sheet_1.max_row) # Print the number of rows in sheet_1
#
# for row in range(2, sheet_1.max_row + 1):
#     cell = sheet_1.cell(row, 3) # Assign cell(row, 3) in sheet_1 to new object cell
#     # print(cell.value)
#     new_price = cell.value * 0.9
#     new_cell = sheet_1.cell(row, 4) # Assign cell(row, 4) in sheet_1 to new object new_cell
#     new_cell.value = new_price  # Assign the value of new_price to value attributes of new_cell
#
#
# values = Reference(sheet_1,
#           min_row=2,
#           max_row=sheet_1.max_row,
#           min_col=4,
#           max_col=4)  # Use Reference class to select a range of values
#
# # print(values)
# # print(values.pop())
#
# chart = BarChart()  # Create a BarChart object called chart
# chart.add_data(values)  # Add the data in Reference object values to chart object
# sheet_1.add_chart(chart, 'E2')  # Add the BarChart object into sheet_1 at cell 'E2'
#
#
# wb.save('transactions2.xlsx')   # Save the changes as a new xlsx file in case errors in code overwrites original xlsx file

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        new_val = 0.9 * cell.value
        new_cell = sheet.cell(row, 4)
        new_cell.value = new_val

    val = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(val)
    sheet.add_chart(chart, 'E2')

    wb.save(filename)


process_workbook('transactions.xlsx')


## Popular Library for Machine Learning
# numpy   # Create multi-dimensional array
# pandas  # Data analysis library that provides concept called data frame (In the nutshell, 2D data structure similar to excel spreadsheet)
# matplotlib  # Plot 2D/3D plots
# scikit-learn # Provides decision tree, neural network algorithms and etc